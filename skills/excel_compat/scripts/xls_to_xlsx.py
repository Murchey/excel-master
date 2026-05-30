import sys
import json
import argparse
import xlrd
import openpyxl
import zipfile
import shutil
from pathlib import Path

def convert_xls_to_xlsx(source_path, target_path=None):
    """
    将 .xls 文件转换为 .xlsx 格式
    
    Args:
        source_path: 源 .xls 文件路径
        target_path: 目标 .xlsx 文件路径（可选）
    
    Returns:
        JSON 格式的执行结果
    """
    source = Path(source_path)
    
    # 防呆校验：文件是否存在
    if not source.exists():
        return json.dumps({
            "status": "error",
            "source_file": str(source),
            "target_file": "",
            "message": f"文件不存在: {source_path}"
        }, ensure_ascii=False)
    
    # 防呆校验：文件扩展名
    if source.suffix.lower() != '.xls':
        return json.dumps({
            "status": "error",
            "source_file": str(source),
            "target_file": "",
            "message": f"仅支持 .xls 格式，当前文件: {source.suffix}"
        }, ensure_ascii=False)
    
    # 确定目标路径
    if target_path:
        target = Path(target_path)
    else:
        target = source.with_suffix('.xlsx')
    
    # 检查目标文件是否已存在
    if target.exists():
        return json.dumps({
            "status": "error",
            "source_file": str(source),
            "target_file": str(target),
            "message": f"目标文件已存在: {target}"
        }, ensure_ascii=False)
    
    try:
        # 检查文件是否是 ZIP 格式（.xlsx 文件实际上是 ZIP 压缩包）
        try:
            with zipfile.ZipFile(str(source), 'r') as zip_ref:
                # 如果是 ZIP 格式，说明文件实际上是 .xlsx 格式
                # 直接复制文件并重命名
                shutil.copy2(str(source), str(target))
                return json.dumps({
                    "status": "success",
                    "source_file": str(source),
                    "target_file": str(target),
                    "message": "文件已是 .xlsx 格式，已复制"
                }, ensure_ascii=False)
        except zipfile.BadZipFile:
            # 不是 ZIP 格式，继续处理真正的 .xls 文件
            pass
        
        # 读取 .xls 文件
        workbook_xls = xlrd.open_workbook(str(source))
        
        # 创建 .xlsx 工作簿
        workbook_xlsx = openpyxl.Workbook()
        
        # 删除默认创建的工作表
        if 'Sheet' in workbook_xlsx.sheetnames:
            del workbook_xlsx['Sheet']
        
        # 遍历所有工作表
        for sheet_name in workbook_xls.sheet_names():
            # 获取源工作表
            sheet_xls = workbook_xls.sheet_by_name(sheet_name)
            
            # 创建目标工作表
            sheet_xlsx = workbook_xlsx.create_sheet(title=sheet_name)
            
            # 复制数据
            for row_idx in range(sheet_xls.nrows):
                for col_idx in range(sheet_xls.ncols):
                    cell_value = sheet_xls.cell_value(row_idx, col_idx)
                    
                    # 处理合并单元格
                    # 注意：xlrd 的 merged_cells 属性返回的是 (row_start, row_end, col_start, col_end) 的列表
                    # 但我们需要检查当前单元格是否在合并范围内
                    
                    # 写入单元格值
                    sheet_xlsx.cell(row=row_idx + 1, column=col_idx + 1, value=cell_value)
            
            # 处理合并单元格
            # xlrd 的 merged_cells 属性在 xlrd 2.0+ 中已被移除
            # 我们需要检查单元格类型来识别合并单元格
            # 由于 xlrd 2.0+ 的变化，我们跳过合并单元格的处理
            # 如果需要保留合并单元格，建议使用其他库如 pandas
        
        # 确保目标目录存在
        target.parent.mkdir(parents=True, exist_ok=True)
        
        # 保存 .xlsx 文件
        workbook_xlsx.save(str(target))
        
        return json.dumps({
            "status": "success",
            "source_file": str(source),
            "target_file": str(target),
            "message": "转换成功"
        }, ensure_ascii=False)
        
    except Exception as e:
        return json.dumps({
            "status": "error",
            "source_file": str(source),
            "target_file": str(target) if target_path else "",
            "message": f"转换失败: {str(e)}"
        }, ensure_ascii=False)

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="将 .xls 文件转换为 .xlsx 格式")
    parser.add_argument("source", help="源 .xls 文件路径")
    parser.add_argument("target", nargs='?', help="目标 .xlsx 文件路径（可选）")
    
    args = parser.parse_args()
    
    # 直接 print 结果，Trae 会捕获 stdout 标准输出
    print(convert_xls_to_xlsx(args.source, args.target))